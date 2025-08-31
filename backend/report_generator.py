import os
import pandas as pd
from openpyxl import load_workbook

REPORTS_DIR = r'Z:/reports'

def yearly_report(classroom, year):
    yearly_file = os.path.join(REPORTS_DIR, f"{classroom}_{year}.xlsx")
    if not os.path.exists(yearly_file):
        print("[ERROR] Yearly Excel file not found.")
        return

    wb = load_workbook(yearly_file)
    summary_data = []
    names_seen = set()

    for sheet in wb.sheetnames:
        if sheet == 'Sheet':  # skip default sheet
            continue
        ws = wb[sheet]
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=2).value
            role = ws.cell(row=row, column=1).value
            if (role, name) in names_seen:
                continue
            names_seen.add((role, name))

            total_classes = 0
            attended = 0
            # Check In columns from 5 to 35 (days 1-31)
            for col in range(5, 5+31):
                if ws.cell(row=row, column=col).value:
                    attended += 1
                total_classes += 1

            percent = round((attended / total_classes) * 100, 2) if total_classes else 0
            summary_data.append([role, name, percent])

    df = pd.DataFrame(summary_data, columns=["Role", "Name", "Attendance %"])
    os.makedirs(os.path.join(REPORTS_DIR, "yearly"), exist_ok=True)
    output_path = os.path.join(REPORTS_DIR, "yearly", f"{classroom}_yearly_{year}.csv")
    df.to_csv(output_path, index=False)
    print(f"[INFO] Yearly attendance summary saved at {output_path}")

if __name__ == "__main__":
    classroom = input("Enter Classroom (e.g. 5EP2): ").upper().strip()
    year = int(input("Enter Year (e.g. 2025): ").strip())
    yearly_report(classroom, year)
