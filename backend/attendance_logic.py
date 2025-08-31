import sys, os, json, pandas as pd
from datetime import datetime
from encryptor import decrypt_embedding

BASE_EMBED_DIR = r'Z:\embeddings'
REPORTS_DIR = r'Z:\reports'
TIMETABLE_FILE = r'Z:\timetable_config.json'

with open(TIMETABLE_FILE, 'r') as f:
    timetable = json.load(f)

# Runtime in-memory session data
session_records = {}

def load_embeddings(classroom):
    folder = os.path.join(BASE_EMBED_DIR, classroom)
    encodings, meta = [], []
    for file in os.listdir(folder):
        if file.endswith(".bin"):
            with open(os.path.join(folder, file), "rb") as f:
                data = decrypt_embedding(f.read())
                encodings.append(data["encoding"])
                meta.append({
                    "name": data["name"],
                    "role": data["role"],
                    "roll": data.get("roll", ""),
                    "classroom": data["classroom"]
                })
    return encodings, meta

def current_session(classroom):
    now = datetime.now().strftime("%H:%M")
    for sess, times in timetable[classroom].items():
        if times["start_time"] <= now <= times["end_time"]:
            return sess
    return None

def mark_attendance(classroom, name, role):
    date_str = datetime.now().strftime("%Y-%m-%d")
    session = current_session(classroom)
    if not session:
        return session, "Invalid Class Timing"

    key = f"{classroom}_{session}_{date_str}"
    if key not in session_records:
        session_records[key] = {"students": {}, "faculty": {}, "faculty_ended": False}

    record_group = "faculty" if role == "faculty" else "students"

    if role == "faculty":
        if name not in session_records[key]["faculty"]:
            session_records[key]["faculty"][name] = {"in": datetime.now().strftime("%H:%M"), "out": ""}
            status = "Present"
        else:
            session_records[key]["faculty"][name]["out"] = datetime.now().strftime("%H:%M")
            session_records[key]["faculty_ended"] = True
            status = "Faculty Marked End"
    else:
        if session_records[key]["faculty_ended"]:
            return session, "Fake Attendance"
        if name not in session_records[key]["students"]:
            session_records[key]["students"][name] = {"in": datetime.now().strftime("%H:%M"), "out": ""}
            status = "Present"
        else:
            session_records[key]["students"][name]["out"] = datetime.now().strftime("%H:%M")
            status = "Out Time Marked"

    log_to_excel(classroom, name, role, session, date_str, session_records[key][record_group][name])
    return session, status

def log_to_excel(classroom, name, role, session, date_str, times):
    year = datetime.now().year
    month = datetime.now().strftime("%B")
    excel_path = os.path.join(REPORTS_DIR, f"{classroom}_{year}.xlsx")

    from openpyxl import Workbook, load_workbook
    try:
        wb = load_workbook(excel_path)
    except:
        wb = Workbook()

    if month not in wb.sheetnames:
        ws = wb.create_sheet(month)
        ws.append(["Role","Name","Roll No","Class"] +
                  [f"Day {d} In" for d in range(1,32)] +
                  [f"Day {d} Out" for d in range(1,32)])
    else:
        ws = wb[month]

    found_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == name and ws.cell(row=row, column=4).value == classroom:
            found_row = row
            break

    if not found_row:
        found_row = ws.max_row + 1
        ws.cell(row=found_row, column=1, value=role)
        ws.cell(row=found_row, column=2, value=name)
        ws.cell(row=found_row, column=3, value="")
        ws.cell(row=found_row, column=4, value=classroom)

    day = int(date_str.split('-')[2])

    ws.cell(row=found_row, column=4 + day, value=times["in"])
    ws.cell(row=found_row, column=4 + 31 + day, value=times["out"])

    wb.save(excel_path)
