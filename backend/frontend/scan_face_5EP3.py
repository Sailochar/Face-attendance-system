import sys
import os
import cv2
import pickle
from datetime import datetime
from cryptography.fernet import Fernet
import json

# SETTINGS
ENCRYPTION_KEY = b'2F4TTtlE05_dM3uo8cD9FpTkqsVycVTfW_HILSowfGA='
TIMETABLE_FILE = r'Z:\timetable_config.json'
BASE_EMBED_DIR = r'Z:\embeddings'
REPORTS_DIR = r'Z:\reports'
CLASSROOM = "5EP3"   # <-- updated for this file!

session_records = {}

def decrypt_embedding(encrypted_bytes):
    fernet = Fernet(ENCRYPTION_KEY)
    pickled = fernet.decrypt(encrypted_bytes)
    return pickle.loads(pickled)

def load_embeddings(classroom):
    folder = os.path.join(BASE_EMBED_DIR, classroom)
    encodings, meta = [], []
    for file in os.listdir(folder):
        if file.endswith(".bin"):
            with open(os.path.join(folder, file), "rb") as f:
                try:
                    data = decrypt_embedding(f.read())
                    encodings.append(data["encoding"])
                    meta.append({
                        "name": data["name"],
                        "role": data["role"],
                        "roll": data.get("roll", ""),
                        "classroom": data["classroom"]
                    })
                except Exception as e:
                    print(f"[ERROR] Failed to decrypt embedding {file}: {e}")
    return encodings, meta

def current_session(classroom, timetable):
    now = datetime.now().strftime("%H:%M")
    for sess, times in timetable.get(classroom, {}).items():
        if times["start_time"] <= now <= times["end_time"]:
            return sess
    return None

def log_to_excel(classroom, name, role, session, date_str, times):
    from openpyxl import Workbook, load_workbook
    year = datetime.now().year
    month = datetime.now().strftime("%B")
    excel_filename = f"{classroom}_{month}_{year}.xlsx"
    excel_path = os.path.join(REPORTS_DIR, excel_filename)

    try:
        wb = load_workbook(excel_path)
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = month
        ws.append(["Role","Name","Roll No","Class"] +
                  [f"Day {d} In" for d in range(1,32)] +
                  [f"Day {d} Out" for d in range(1,32)])

    if month in wb.sheetnames:
        ws = wb[month]
    else:
        ws = wb.create_sheet(month)
        ws.append(["Role","Name","Roll No","Class"] +
                  [f"Day {d} In" for d in range(1,32)] +
                  [f"Day {d} Out" for d in range(1,32)])

    found_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == name and ws.cell(row=row, column=4).value == classroom:
            found_row = row
            break
    if not found_row:
        found_row = ws.max_row + 1
        ws.cell(row=found_row, column=1, value=role)
        ws.cell(row=found_row, column=2, value=name)
        ws.cell(row=found_row, column=3, value="")
        ws.cell(row=found_row, column=4, value=classroom)

    day = int(date_str.split('-')[2])
    ws.cell(row=found_row, column=4 + day, value=times["in"])
    ws.cell(row=found_row, column=4 + 31 + day, value=times["out"])

    wb.save(excel_path)

def mark_attendance(classroom, name, role, timetable):
    date_str = datetime.now().strftime("%Y-%m-%d")
    session = current_session(classroom, timetable)
    if not session:
        return session, "Invalid Class Timing"
    key = f"{classroom}_{session}_{date_str}"
    if key not in session_records:
        session_records[key] = {"students": {}, "faculty": {}, "faculty_ended": False}
    record_group = "faculty" if role == "faculty" else "students"
    if role == "faculty":
        if name not in session_records[key]["faculty"]:
            session_records[key]["faculty"][name] = {"in": datetime.now().strftime("%H:%M"), "out": ""}
            status = "Present"
        else:
            session_records[key]["faculty"][name]["out"] = datetime.now().strftime("%H:%M")
            session_records[key]["faculty_ended"] = True
            status = "Faculty Marked End"
    else:
        if session_records[key]["faculty_ended"]:
            return session, "Fake Attendance"
        if name not in session_records[key]["students"]:
            session_records[key]["students"][name] = {"in": datetime.now().strftime("%H:%M"), "out": ""}
            status = "Present"
        else:
            session_records[key]["students"][name]["out"] = datetime.now().strftime("%H:%M")
            status = "Out Time Marked"
    log_to_excel(classroom, name, role, session, date_str, session_records[key][record_group][name])
    return session, status

def draw_panel(frame, classroom, session, status, name):
    panel_height = 60
    cv2.rectangle(frame, (0, frame.shape[0] - panel_height), (frame.shape, frame.shape), (30, 30, 30), -1)
    info = f"Classroom: {classroom} | Session: {session if session else 'No Class'} | Time: {datetime.now().strftime('%H:%M:%S')}"
    cv2.putText(frame, info, (10, frame.shape[0] - panel_height + 25), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)
    scan_info = f"Last Scan: {name if name else '-'} | Status: {status if status else '-'}"
    if status is None or status == "Invalid Class Timing":
        color = (50, 50, 255)
    elif "Present" in status:
        color = (0, 255, 0)
    elif "Half" in status:
        color = (0, 255, 255)
    else:
        color = (0, 0, 255)
    cv2.putText(frame, scan_info, (10, frame.shape[0] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)

def main():
    import face_recognition
    with open(TIMETABLE_FILE, 'r') as f:
        timetable = json.load(f)
    os.makedirs(REPORTS_DIR, exist_ok=True)
    encodings, meta = load_embeddings(CLASSROOM)
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("[ERROR] Cannot access camera.")
        return
    last_name, last_status = "", "Waiting for scan..."
    while True:
        ret, frame = cap.read()
        if not ret:
            print("[ERROR] Frame capture failed.")
            break
        session = current_session(CLASSROOM, timetable)
        if session and encodings:
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            face_locations = face_recognition.face_locations(rgb_frame)
            face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                matches = face_recognition.compare_faces(encodings, face_encoding, tolerance=0.5)
                if True in matches:
                    idx = matches.index(True)
                    name, role = meta[idx]["name"], meta[idx]["role"]
                    session_name, status = mark_attendance(CLASSROOM, name, role, timetable)
                    last_name, last_status = name, status
                    if "Present" in status:
                        color = (0, 255, 0)
                    elif "Half" in status:
                        color = (0, 255, 255)
                    else:
                        color = (0, 0, 255)
                    cv2.rectangle(frame, (left, top), (right, bottom), color, 3)
                    cv2.putText(frame, f"{name} | {status}", (left, top-12), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
                else:
                    last_name, last_status = "", "Unknown"
                    cv2.rectangle(frame, (left, top), (right, bottom), (0,0,255), 2)
                    cv2.putText(frame, "Unknown", (left, top-12), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,0,255), 2)
        else:
            last_name, last_status = "", "Invalid Class Timing"
        draw_panel(frame, CLASSROOM, session, last_status, last_name)
        cv2.imshow(f"{CLASSROOM} Attendance Scanner", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    main()
